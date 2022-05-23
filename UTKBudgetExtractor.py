#!/usr/bin/env python

import sys,os
import math
from datetime import datetime

from openpyxl import load_workbook
from collections import OrderedDict

# Give the location of the file at the command line

loc = sys.argv[1]

# Load the spreadsheet file in xlsx format
# `data_only=True` forces it to calculate all formulas

wb = load_workbook(filename = loc, data_only=True)
ws = wb.active

commandDict = OrderedDict()

def abc(i):
    return chr(ord('@')+i+1)

# Section A

for iSenior in range(10):
    commandDict[f"Senior{abc(iSenior)}Name"]      = ws[f"B{iSenior+11}"].value
    commandDict[f"Senior{abc(iSenior)}BaseAnnual"]= ws[f"D{iSenior+11}"].value
    commandDict[f"Senior{abc(iSenior)}YearOne"]   = ws[f"L{iSenior+11}"].value
    commandDict[f"Senior{abc(iSenior)}YearTwo"]   = ws[f"M{iSenior+11}"].value
    commandDict[f"Senior{abc(iSenior)}YearThree"] = ws[f"N{iSenior+11}"].value
    commandDict[f"Senior{abc(iSenior)}YearFour"]  = ws[f"O{iSenior+11}"].value
    commandDict[f"Senior{abc(iSenior)}YearFive"]  = ws[f"P{iSenior+11}"].value
    commandDict[f"Senior{abc(iSenior)}Total"]     = ws[f"Q{iSenior+11}"].value

commandDict[f"SeniorTotal"] = ws[f"Q21"].value

# Section B

for iPostdoc in range(3):
    commandDict[f"Postdoc{abc(iPostdoc)}Name"]           = ws[f"B{iPostdoc+24}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}BaseMonthly"]    = ws[f"D{iPostdoc+24}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}YearOne"]        = ws[f"L{iPostdoc+24}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}YearTwo"]        = ws[f"M{iPostdoc+24}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}YearThree"]      = ws[f"N{iPostdoc+24}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}YearFour"]       = ws[f"O{iPostdoc+24}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}YearFive"]       = ws[f"P{iPostdoc+24}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}Total"]          = ws[f"Q{iPostdoc+24}"].value


for iOther in range(3):
    commandDict[f"Other{abc(iOther)}Name"]               = ws[f"B{iOther+27}"].value
    commandDict[f"Other{abc(iOther)}BaseMonthly"]        = ws[f"D{iOther+27}"].value
    commandDict[f"Other{abc(iOther)}YearOne"]            = ws[f"L{iOther+27}"].value
    commandDict[f"Other{abc(iOther)}YearTwo"]            = ws[f"M{iOther+27}"].value
    commandDict[f"Other{abc(iOther)}YearThree"]          = ws[f"N{iOther+27}"].value
    commandDict[f"Other{abc(iOther)}YearFour"]           = ws[f"O{iOther+27}"].value
    commandDict[f"Other{abc(iOther)}YearFive"]           = ws[f"P{iOther+27}"].value
    commandDict[f"Other{abc(iOther)}Total"]              = ws[f"Q{iOther+27}"].value


for iGRA in range(4):
    commandDict[f"GRA{abc(iGRA)}Name"]                   = ws[f"B{iGRA+30}"].value
    commandDict[f"GRA{abc(iGRA)}BaseMonthly"]            = ws[f"D{iGRA+30}"].value
    commandDict[f"GRA{abc(iGRA)}YearOne"]                = ws[f"L{iGRA+30}"].value
    commandDict[f"GRA{abc(iGRA)}YearTwo"]                = ws[f"M{iGRA+30}"].value
    commandDict[f"GRA{abc(iGRA)}YearThree"]              = ws[f"N{iGRA+30}"].value
    commandDict[f"GRA{abc(iGRA)}YearFour"]               = ws[f"O{iGRA+30}"].value
    commandDict[f"GRA{abc(iGRA)}YearFive"]               = ws[f"P{iGRA+30}"].value
    commandDict[f"GRA{abc(iGRA)}Total"]                  = ws[f"Q{iGRA+30}"].value

for iURA in range(1):
    commandDict[f"UGA{abc(iURA)}Name"]                   = ws[f"B{iURA+34}"].value
    commandDict[f"UGA{abc(iURA)}BaseMonthly"]            = ws[f"D{iURA+34}"].value
    commandDict[f"UGA{abc(iURA)}YearOne"]                = ws[f"L{iURA+34}"].value
    commandDict[f"UGA{abc(iURA)}YearTwo"]                = ws[f"M{iURA+34}"].value
    commandDict[f"UGA{abc(iURA)}YearThree"]              = ws[f"N{iURA+34}"].value
    commandDict[f"UGA{abc(iURA)}YearFour"]               = ws[f"O{iURA+34}"].value
    commandDict[f"UGA{abc(iURA)}YearFive"]               = ws[f"P{iURA+34}"].value
    commandDict[f"UGA{abc(iURA)}Total"]                  = ws[f"Q{iURA+34}"].value


for iSecretary in range(1):
    commandDict[f"Secretary{abc(iSecretary)}Name"]       = ws[f"B{iSecretary+35}"].value
    commandDict[f"Secretary{abc(iSecretary)}BaseMonthly"]= ws[f"D{iSecretary+35}"].value
    commandDict[f"Secretary{abc(iSecretary)}YearOne"]    = ws[f"L{iSecretary+35}"].value
    commandDict[f"Secretary{abc(iSecretary)}YearTwo"]    = ws[f"M{iSecretary+35}"].value
    commandDict[f"Secretary{abc(iSecretary)}YearThree"]  = ws[f"N{iSecretary+35}"].value
    commandDict[f"Secretary{abc(iSecretary)}YearFour"]   = ws[f"O{iSecretary+35}"].value
    commandDict[f"Secretary{abc(iSecretary)}YearFive"]   = ws[f"P{iSecretary+35}"].value
    commandDict[f"Secretary{abc(iSecretary)}Total"]      = ws[f"Q{iSecretary+35}"].value

commandDict[f"OtherTotal"] = ws[f"Q37"].value
commandDict[f"WagesTotal"] = ws[f"Q38"].value

# Section C


for iSenior in range(10):
    commandDict[f"Senior{abc(iSenior)}FringeName"]       = ws[f"B{iSenior+41}"].value
    commandDict[f"Senior{abc(iSenior)}FringeRate"]       = ws[f"F{iSenior+41}"].value * 100 #percent
    commandDict[f"Senior{abc(iSenior)}FringeYearOne"]    = ws[f"L{iSenior+41}"].value
    commandDict[f"Senior{abc(iSenior)}FringeYearTwo"]    = ws[f"M{iSenior+41}"].value
    commandDict[f"Senior{abc(iSenior)}FringeYearThree"]  = ws[f"N{iSenior+41}"].value
    commandDict[f"Senior{abc(iSenior)}FringeYearFour"]   = ws[f"O{iSenior+41}"].value
    commandDict[f"Senior{abc(iSenior)}FringeYearFive"]   = ws[f"P{iSenior+41}"].value
    commandDict[f"Senior{abc(iSenior)}FringeTotal"]      = ws[f"Q{iSenior+41}"].value

for iPostdoc in range(3):
    commandDict[f"Postdoc{abc(iPostdoc)}FringeName"]     = ws[f"B{iPostdoc+51}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}FringeRate"]     = ws[f"F{iPostdoc+51}"].value * 100 #percent
    commandDict[f"Postdoc{abc(iPostdoc)}FringeYearOne"]  = ws[f"L{iPostdoc+51}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}FringeYearTwo"]  = ws[f"M{iPostdoc+51}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}FringeYearThree"]= ws[f"N{iPostdoc+51}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}FringeYearFour"] = ws[f"O{iPostdoc+51}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}FringeYearFive"] = ws[f"P{iPostdoc+51}"].value
    commandDict[f"Postdoc{abc(iPostdoc)}FringeTotal"]    = ws[f"Q{iPostdoc+51}"].value


commandDict[f"GRAHealthName"]          = ws[f"B61"].value
commandDict[f"GRAHealthRate"]          = ws[f"F61"].value
commandDict[f"GRAHealthYearOne"]       = ws[f"L61"].value
commandDict[f"GRAHealthYearTwo"]       = ws[f"M61"].value
commandDict[f"GRAHealthYearThree"]     = ws[f"N61"].value
commandDict[f"GRAHealthYearFour"]      = ws[f"O61"].value
commandDict[f"GRAHealthYearFive"]      = ws[f"P61"].value
commandDict[f"GRAHealthTotal"]         = ws[f"Q61"].value

commandDict[f"FringeTotal"]             = ws[f"Q62"].value
commandDict[f"SalaryAndBenefitsTotal"]  = ws[f"Q63"].value

# Section E

commandDict[f"DomesticTravelTotal"]     = ws[f"Q73"].value
commandDict[f"InternationalTravelTotal"]= ws[f"Q74"].value
commandDict[f"TravelTotal"]             = ws[f"Q75"].value

# Section G

commandDict[f"SuppliesTotal"]           = ws[f"Q80"].value

commandDict[f"TuitionYearOne"]          = ws[f"L91"].value
commandDict[f"TuitionYearTwo"]          = ws[f"M91"].value
commandDict[f"TuitionYearThree"]        = ws[f"N91"].value
commandDict[f"TuitionYearFour"]         = ws[f"O91"].value
commandDict[f"TuitionYearFive"]         = ws[f"P91"].value
commandDict[f"TuitionTotal"]            = ws[f"Q91"].value

commandDict[f"OtherDirectTotal"]        = ws[f"Q92"].value

# Section H

commandDict[f"DirectTotal"]             = ws[f"Q94"].value

# Section I

commandDict[f"OverheadRate"]            = ws[f"L98"].value * 100 #percent

# Section J

commandDict[f"IndirectTotal"]           = ws[f"Q99"].value

# Section K

commandDict[f"Total"]                   = ws[f"Q100"].value


with open('budgetDefs.tex', 'w') as outputFile:

    outputFile.write(f'% {datetime.now()}' + os.linesep)
    outputFile.write(f'% Auto-generated by {sys.argv[0]}' + os.linesep)
    outputFile.write(f'% >>> {" ".join(sys.argv)}' + os.linesep)
    outputFile.write(f'% >>> Written by L Lee for UTK Budget Spreadsheet as of 2022' + os.linesep)
    outputFile.write(os.linesep)


    for key,val in commandDict.items():
        if not val:
            val = ""
        try:
            if math.modf(float(val))[0]:
                val = "%.2f" % val
        except:
            pass
        
        print("{: >30} {: >20}".format(key,val))

        outputFile.write("\\newcommand{\\%s}{%s}"%(key,val) + os.linesep)

print(">>>")
print(">>> Write output to budgetDefs.tex!")

